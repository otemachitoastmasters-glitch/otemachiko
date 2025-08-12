import os
import re
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook

# 画像は存在する場合のみ挿入（Pillow 必須）
try:
    from openpyxl.drawing.image import Image
    PIL_OK = True
except Exception:
    PIL_OK = False


# -----------------------------
# 共通: HTTPヘッダ（UA）
# -----------------------------
UA_DESKTOP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0 Safari/537.36"
    )
}
UA_MOBILE = {
    "User-Agent": (
        "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
        "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 "
        "Mobile/15E148 Safari/604.1"
    )
}


# -----------------------------
# 1) <a onclick="showDetail(n)"> の最初の n を返す
#    （PC→モバイルの順で試す）
# -----------------------------
def fetch_first_mtgid_by_showdetail(url="https://tmcsupport.coresv.com/otemachiko/"):
    for headers in (UA_DESKTOP, UA_MOBILE):
        try:
            r = requests.get(url, headers=headers, timeout=15)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            for a_tag in soup.find_all("a", onclick=True):
                m = re.search(r"showDetail\((\d+)\)", a_tag["onclick"])
                if m:
                    return int(m.group(1))
        except Exception:
            continue
    return None


# -----------------------------
# 2) テーブル（PC表示）から「今日以降で最も近い」 mtgid
# -----------------------------
def fetch_latest_mtgid(base_url="https://tmcsupport.coresv.com/otemachiko/"):
    try:
        resp = requests.get(base_url, headers=UA_DESKTOP, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        rows = soup.select("table.tableCommon tr")[1:]  # ヘッダ行除外
        latest_mtgid = None
        latest_date = None
        today = datetime.today()

        for row in rows:
            tds = row.find_all("td")
            if len(tds) < 2:
                continue
            date_text = tds[0].get_text(strip=True)
            a = tds[1].find("a")
            if not a or "onclick" not in a.attrs:
                continue
            m = re.search(r"showDetail\((\d+)\)", a["onclick"])
            if not m:
                continue

            try:
                d = datetime.strptime(date_text, "%Y/%m/%d")
            except Exception:
                continue

            if d >= today and (latest_date is None or d < latest_date):
                latest_date = d
                latest_mtgid = int(m.group(1))

        return latest_mtgid
    except Exception:
        return None


# -----------------------------
# 小物: 安全書き込み＆タイトル括弧付け
# -----------------------------
def safe_set(ws, cell, value):
    """値があれば書く。なければ空文字を書いておく（落ちない）"""
    ws[cell] = value if value else ""

def safe_quote(value, left="「", right="」"):
    return f"{left}{value}{right}" if value else ""


# -----------------------------
# 3) Excel生成（mtgid 指定）
# -----------------------------
def generate_agenda_excel_from_url(
    mtgid: int,
    template_path: str = "meeting_agenda_template.xlsx",
    output_path: str = "generated_agenda.xlsx",
) -> str:

    html_url = f"https://tmcsupport.coresv.com/otemachiko/mtgDetailReadonly.php?mtgid={mtgid}"
    print(f"🔗 Fetching agenda from: {html_url}")

    # HTML 取得（PC UA）
    res = requests.get(html_url, headers=UA_DESKTOP, timeout=20)
    res.raise_for_status()
    soup = BeautifulSoup(res.content, "html.parser")

    # --- 会議情報 ---
    header_table = soup.find("table", class_="tableCommon")
    if not header_table:
        raise RuntimeError("Header table (tableCommon) not found.")

    rows = header_table.find_all("tr")
    if len(rows) < 2:
        raise RuntimeError("Header table has no data row.")

    mtg_info = rows[1].find_all("td")
    # ガード（列不足に備える）
    def td_text(idx):
        return mtg_info[idx].get_text(strip=True) if len(mtg_info) > idx else ""

    date = td_text(0)
    title = td_text(1)
    venue = td_text(3)
    room = td_text(4)

    meeting_title = title
    meeting_datetime = date

    # --- Guests ---
    guest = ""
    for table in soup.find_all("table", class_="tableCommon"):
        th = table.find("th")
        if th and "Guests" in th.get_text():
            td = table.find("td")
            guest = td.get_text(strip=True) if td else ""
            break

    # --- アジェンダ表 ---
    agenda_table = soup.find("table", class_="tableCommon mainTbl")
    if not agenda_table:
        raise RuntimeError("Agenda table (tableCommon mainTbl) not found.")

    role_name_map = {}
    evaluator_map = {}
    speech_path_map = {}
    speech_title_map = {}
    theme = ""

    for tr in agenda_table.find_all("tr")[1:]:
        tds = tr.find_all("td")
        if len(tds) < 2:
            continue
        role = tds[0].get_text(strip=True)
        name = tds[1].get_text(strip=True) if len(tds) > 1 else ""
        detail = tds[2].get_text("\n", strip=True) if len(tds) > 2 else ""
        title_cell = tds[3].get_text(strip=True) if len(tds) > 3 else ""

        # 役割→名前
        role_name_map[role] = name

        # テーマ
        if "Theme" in role:
            theme = detail

        # スピーチ（Path/Title抽出）
        if role.startswith("Speech"):
            # detail は複数行の可能性があるので最後の行など必要に応じて調整
            # 例: PathやLevel/Projectが含まれるブロック → ここでは丸ごと入れる
            speech_path_map[role] = detail or ""
            speech_title_map[role] = title_cell or ""

        # Evaluator
        if role.startswith("Evaluator"):
            evaluator_map[role] = detail or ""

    # --- Excel ---
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active
    ws.title = "Agenda"

    # 画像（存在する場合のみ）
    logo_path = "toastmasters_logo.jpg"
    if PIL_OK and os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            img.width = 150
            img.height = 105
            ws.add_image(img, "B1")
        except Exception:
            pass  # 画像は任意

    # すべての結合セルを一旦解除（書き込みエラー回避）
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    # ヘッダ
    safe_set(ws, "F3", theme)
    safe_set(ws, "K3", f"{meeting_title}　{meeting_datetime}")
    safe_set(ws, "K4", f"{venue} {room}")

    # 役割：存在しない場合は空でOK
    safe_set(ws, "I9",  role_name_map.get("Toastmaster of the Evening"))
    safe_set(ws, "I10", role_name_map.get("Word of the Evening"))
    safe_set(ws, "I11", role_name_map.get("Ah-Counter"))
    safe_set(ws, "I12", role_name_map.get("Grammarian"))
    safe_set(ws, "I13", role_name_map.get("Timer"))
    safe_set(ws, "I14", role_name_map.get("PC Manager (Vote Counter)"))

    # Table Topics / Prepared Speech
    safe_set(ws, "I16", role_name_map.get("Table Topics Master"))

    # Speech1..4（なければスキップ）
    safe_set(ws, "E25", safe_quote(speech_title_map.get("Speech1")))
    safe_set(ws, "I25", speech_path_map.get("Speech1"))
    safe_set(ws, "I26", role_name_map.get("Speech1"))

    safe_set(ws, "E27", safe_quote(speech_title_map.get("Speech2")))
    safe_set(ws, "I27", speech_path_map.get("Speech2"))
    safe_set(ws, "I28", role_name_map.get("Speech2"))

    safe_set(ws, "E29", safe_quote(speech_title_map.get("Speech3")))
    safe_set(ws, "I29", speech_path_map.get("Speech3"))
    safe_set(ws, "I30", role_name_map.get("Speech3"))

    safe_set(ws, "E31", safe_quote(speech_title_map.get("Speech4")))
    safe_set(ws, "I31", speech_path_map.get("Speech4"))
    safe_set(ws, "I32", role_name_map.get("Speech4"))

    # GE, Evaluators（存在しないキーは空）
    safe_set(ws, "I37", role_name_map.get("General Evaluator"))

    safe_set(ws, "E38", evaluator_map.get("Evaluator1"))
    safe_set(ws, "I38", role_name_map.get("Evaluator1"))

    safe_set(ws, "E39", evaluator_map.get("Evaluator2"))
    safe_set(ws, "I39", role_name_map.get("Evaluator2"))

    safe_set(ws, "E40", evaluator_map.get("Evaluator3"))
    safe_set(ws, "I40", role_name_map.get("Evaluator3"))

    safe_set(ws, "E41", evaluator_map.get("Evaluator4"))
    safe_set(ws, "I41", role_name_map.get("Evaluator4"))

    # レポート欄（Woe/Ah/Grammar）
    safe_set(ws, "I43", role_name_map.get("Word of the Evening"))
    safe_set(ws, "I44", role_name_map.get("Ah-Counter"))
    safe_set(ws, "I45", role_name_map.get("Grammarian"))

    # 保存
    out = output_path or f"{meeting_title or 'Agenda'}_agenda.xlsx"
    wb.save(out)
    print(f"✅ Saved Excel to: {out}")
    return out
